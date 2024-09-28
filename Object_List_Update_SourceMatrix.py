import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def load_excel_data(input_file, sheet_name):
    """Load Excel file and return data as DataFrame"""
    try:
        xls = pd.ExcelFile(input_file)
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        return df, xls
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return None, None
    except ValueError as e:
        print(f"Error: {e}")
        return None, None

def find_headers(object_list_df):
    """Find the header rows for Data Warehouse and Data Mart sections."""
    try:
        dwh_header_row = 0  # Assuming first row is always Data Warehouse header
        data_mart_header_row = object_list_df[object_list_df.isnull().all(axis=1)].index[0] + 1
        return dwh_header_row, data_mart_header_row
    except IndexError:
        print("Error: Could not locate Data Mart header row.")
        return None, None

def read_data_tables(xls, dwh_header_row, data_mart_header_row):
    """Read Data Warehouse and Data Mart tables from the 'object list' sheet."""
    table1_df = pd.read_excel(xls, sheet_name='object list', header=dwh_header_row, nrows=data_mart_header_row - dwh_header_row - 1).dropna(how='all')
    table2_df = pd.read_excel(xls, sheet_name='object list', header=data_mart_header_row).dropna(how='all')
    return table1_df, table2_df

def fetch_description_from_sheet(wb, sheet_name):
    """Fetch the description from Cell B5 of the given worksheet."""
    try:
        sheet = wb[sheet_name]
        description = sheet["B5"].value  # Fetching the value from cell B5
        if description:
            return description
        else:
            return None
    except KeyError:
        return None

def process_new_objects(sheet_names, existing_dwh_objects, existing_dm_objects, wb):
    """Find new objects in sheet names that are not already in the object list."""
    new_dwh_objects, new_dm_objects = [], []
    for sheet_name in sheet_names:
        cleaned_sheet_name = sheet_name.strip().lower()
        description = fetch_description_from_sheet(wb, sheet_name)  # Fetch description from B5
        if cleaned_sheet_name.startswith('dwh.') and sheet_name not in existing_dwh_objects:
            new_dwh_objects.append([None, sheet_name, description if description else 'New description', '', ''])
        elif cleaned_sheet_name.startswith('dm.') and sheet_name not in existing_dm_objects:
            new_dm_objects.append([None, sheet_name, description if description else 'New description', ''])
    return new_dwh_objects, new_dm_objects

def create_and_format_workbook(output_file, updated_table1_df, updated_table2_df):
    """Create new workbook and write data with formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'object list'

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    hyperlink_font = Font(underline='single', color='0000FF')

    # Write Data Warehouse headers and format
    dwh_headers = ['Dwh object number', 'Dwh object name', 'Dwh object description', 'Reports Tag', 'checked']
    for col_idx, header in enumerate(dwh_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)  # Bold and font size +2
        cell.border = thin_border

    # Write Data Warehouse objects
    for r_idx, row in enumerate(updated_table1_df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            # Alignments and hyperlinking
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = hyperlink_font
                cell.hyperlink = f"#{row[1]}!A1"  # Correct format for internal link
            if c_idx == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if row[2] != 'New description':  # Apply bold only if it's not the default 'New description'
                    cell.font = bold_font
                ws.row_dimensions[r_idx].height = 45

    # Write Data Mart headers and objects
    dm_start_row = len(updated_table1_df) + 3
    dm_headers = ['Data mart object number', 'Data mart object name', 'Data mart object description', 'Reports Tag']
    for col_idx, header in enumerate(dm_headers, 1):
        cell = ws.cell(row=dm_start_row - 1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=12)  # Bold and font size +2
        cell.border = thin_border

    for r_idx, row in enumerate(updated_table2_df.itertuples(index=False), dm_start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            if c_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = hyperlink_font
                cell.hyperlink = f"#{row[1]}!A1"  # Correct format for internal link
            if c_idx == 3:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if row[2] != 'New description':  # Apply bold only if it's not the default 'New description'
                    cell.font = bold_font
                ws.row_dimensions[r_idx].height = 45

    # Adjust column widths
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 160  # 8x width of column B
    ws.column_dimensions['A'].width = 15

    # Save the new workbook
    wb.save(output_file)
    print(f"New version of Excel file created and saved to {output_file}")

def main():
    input_file = r"C:\Users\pttom\OneDrive\Pulpit\DWH_Source_Matrix (2).xlsx"
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = rf"C:\Users\pttom\OneDrive\Pulpit\Object_List_Only_{timestamp}.xlsx"

    # Load data and sheet names
    object_list_df, xls = load_excel_data(input_file, 'object list')
    if object_list_df is None or xls is None:
        return  # Exit if error in loading

    # Load the workbook to access other sheets
    wb = openpyxl.load_workbook(input_file)

    # Find headers for DWH and Data Mart sections
    dwh_header_row, data_mart_header_row = find_headers(object_list_df)
    if dwh_header_row is None or data_mart_header_row is None:
        return  # Exit if error in finding headers

    # Read tables and existing objects
    table1_df, table2_df = read_data_tables(xls, dwh_header_row, data_mart_header_row)
    existing_dwh_objects = table1_df['Dwh object name'].tolist()
    existing_dm_objects = table2_df['Data mart object name'].tolist()

    # Find new objects
    new_dwh_objects, new_dm_objects = process_new_objects(xls.sheet_names, existing_dwh_objects, existing_dm_objects, wb)

    # Convert new objects to DataFrames and concatenate
    new_dwh_df = pd.DataFrame(new_dwh_objects, columns=['Dwh object number', 'Dwh object name', 'Dwh object description', 'Reports Tag', 'checked'])
    new_dm_df = pd.DataFrame(new_dm_objects, columns=['Data mart object number', 'Data mart object name', 'Data mart object description', 'Reports Tag'])
    updated_table1_df = pd.concat([table1_df, new_dwh_df], ignore_index=True).sort_values(by='Dwh object name').reset_index(drop=True)
    updated_table2_df = pd.concat([table2_df, new_dm_df], ignore_index=True).sort_values(by='Data mart object name').reset_index(drop=True)

    # Renumber the objects
    updated_table1_df['Dwh object number'] = updated_table1_df.index + 1
    updated_table2_df['Data mart object number'] = updated_table2_df.index + 1

    # Create and save new Excel file
    create_and_format_workbook(output_file, updated_table1_df, updated_table2_df)

if __name__ == "__main__":
    main()
