# ------------------------------------------------------------------------------
# Version: 1.6
# Last updated: 2024-09-12
# Description: 
# This script extracts entity column names and PK/FK information from an Excel 
# file, applies formatting such as color-coding for different systems and PK/FK 
# indicators, and saves the results to a new Excel file. Additionally, it creates 
# a legend sheet with color codes for system identifiers and FK color. It also 
# includes ASCII art for fun at the end.
# ------------------------------------------------------------------------------

import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Path to the source Excel file (use raw string)
source_file = r'c:\\python\\DWH_Source_Matrix (2).xlsx'

# Get the current timestamp in the format YYYYMMDD_HHMMSS
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

# Define the output file path with the timestamp appended to the file name (use raw string)
output_file = rf'c:\\python\\DWH_Entity_Columns_Output_{timestamp}.xlsx'

# Load the Excel file
excel_file = pd.ExcelFile(source_file)

# Ensure the color is in aRGB format (prepend 'FF' for fully opaque colors)
def convert_to_argb(rgb_hex):
    return f'FF{rgb_hex}' if len(rgb_hex) == 6 else rgb_hex

# Define a dictionary that maps system identifiers to colors (converted to aRGB)
system_color_map = {
    '01': ('D365', convert_to_argb('D9E1F2')),    # Light Blue for D365
    '02': ('AXAPTA', convert_to_argb('C6E0B4')),  # Light Green for AXAPTA
    '03': ('SalesForce', convert_to_argb('FFF2CC')),  # Light Yellow for SalesForce
    '04': ('PDM', convert_to_argb('F4CCCC')),     # Light Pink for PDM
    '05': ('Mobile Installer', convert_to_argb('D9D9D9')),  # Light Gray for Mobile Installer
    '06': ('Order Management', convert_to_argb('FFD966')),  # Light Orange for OM
    '07': ('DWH', convert_to_argb('EAD1DC')),     # Light Purple for DWH
    '08': ('Budget.xlsx', convert_to_argb('D0E0E3')),  # Light Teal for Budget.xlsx
    '09': ('AccountItemMap.xlsx', convert_to_argb('F4CCCC'))  # Light Red for AccountItemMap.xlsx
}

# Initialize a dictionary to store the extracted columns and PK/FK info for each sheet
extracted_columns = {}
pk_fk_info_columns = {}

# Loop through each sheet to find 'entity column name' and extract additional info, including system identifier
for sheet in excel_file.sheet_names:
    df = pd.read_excel(source_file, sheet_name=sheet)

    if df.shape[1] > 6:
        # Look for the cell that contains 'entity column name' (case insensitive) in column D
        entity_row = df[df.iloc[:, 3].str.contains('entity column name', case=False, na=False)]

        if not entity_row.empty:
            start_index = entity_row.index[0] + 1

            entity_columns = []
            pk_fk_info = []
            for idx, value in df.iloc[start_index:, 3].items():
                if pd.isna(value):  # Stop when an empty cell is encountered
                    break
                entity_columns.append(value)

                description = df.iloc[idx, 6]  # Column G for description
                source_system = 'DWH'  # Default to DWH (light purple) for unrecognized systems
                color = convert_to_argb('EAD1DC')  # Default color for DWH (light purple)

                # If a PK or FK is found, search column A for the matching entity column
                if not pd.isna(description) and ('primary key' in description.lower() or 'unique key' in description.lower()):
                    # Search for the matching entity column name in column A
                    entity_name = value
                    match_row = None
                    for search_idx, search_value in df.iloc[:, 0].items():
                        if search_value == entity_name:
                            match_row = search_idx
                            break

                    if match_row is not None and match_row < len(df) - 1:
                        # Look in the row below for the source system identifier in column A
                        column_a_value = str(df.iloc[match_row + 1, 0])
                        if len(column_a_value) >= 2:
                            source_system_id = column_a_value[-2:]  # Extract the last two characters
                            if source_system_id in system_color_map:
                                source_system, color = system_color_map[source_system_id]

                # Initialize the dictionary to hold information for each column
                column_info = {
                    'type': '',
                    'custom_identifier': '',
                    'description': str(description) if not pd.isna(description) else '',
                    'source_system': source_system,
                    'color': color
                }

                # Determine PK or FK status with custom identifier
                if pd.isna(description):
                    column_info['type'] = ''  # Neither PK nor FK
                elif re.search(r'\bPK\b', str(description), re.IGNORECASE) or \
                        any(kw in str(description).lower() for kw in ['primary key', 'unique identifier', 'unique key']):
                    column_info['type'] = 'PK'
                    column_info['custom_identifier'] = f'PK{len(pk_fk_info) + 1:02d}'  # Assign 'PK01', 'PK02', etc.
                elif 'fk' in str(description).lower() or 'foreign key' in str(description).lower():
                    column_info['type'] = 'FK'
                    column_info['custom_identifier'] = f'FK{len(pk_fk_info) + 1:02d}'  # Assign 'FK01', 'FK02', etc.

                # Add the column info to the pk_fk_info list for this sheet
                pk_fk_info.append(column_info)

            # Store the extracted entity column names and PK/FK info for the current sheet
            extracted_columns[sheet] = entity_columns
            pk_fk_info_columns[sheet] = pk_fk_info

# Create a DataFrame for Sheet 1
extracted_df = pd.DataFrame.from_dict(extracted_columns, orient='index').transpose()

# Save the extracted columns to a new Excel file (Sheet 1)
with pd.ExcelWriter(output_file) as writer:
    extracted_df.to_excel(writer, sheet_name='Extracted Columns', index=False)

# Now, open the newly created Excel file for coloring and modifications
wb = load_workbook(output_file)
ws = wb['Extracted Columns']  # Open Sheet 1 for coloring

# Define new fill colors for FK and system colors for PK
fk_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Light gray for FK
red_font = Font(color="FF0000", bold=True)  # Red font for the sheet name if multiple PKs
pk_bold_font = Font(bold=True)  # Bold font for PK

# Track the number of PKs per sheet
pk_count = {}

# Loop through the extracted columns and apply coloring based on PK/FK detection
for sheet_name, entity_columns in extracted_columns.items():
    pk_fk_info = pk_fk_info_columns[sheet_name]  # Get corresponding PK/FK info for this sheet

    # Add hyperlinks for the sheet name
    hyperlink_formula = f'=HYPERLINK("#\'{sheet_name}\'!A1", "{sheet_name}")'
    ws.cell(row=1, column=list(extracted_columns.keys()).index(sheet_name) + 1, value=hyperlink_formula)

    for idx, entity_column in enumerate(entity_columns):
        # Get the PK/FK info for the corresponding entity column
        pk_fk = pk_fk_info[idx]

        if pk_fk['type'] == 'PK':  # If it's a Primary Key
            # Apply the color specific to the PK based on its source system
            if len(pk_fk['color']) == 8:  # Ensure color is valid aRGB hex
                cell = ws.cell(row=idx + 2, column=list(extracted_columns.keys()).index(sheet_name) + 1)
                cell.fill = PatternFill(start_color=pk_fk['color'], end_color=pk_fk['color'], fill_type="solid")
                cell.font = pk_bold_font  # Make PK bold
        
        elif pk_fk['type'] == 'FK':  # If it's a Foreign Key
            # Apply the fixed FK color
            cell = ws.cell(row=idx + 2, column=list(extracted_columns.keys()).index(sheet_name) + 1)
            cell.fill = fk_fill

# Create a new sheet for the legend
legend_ws = wb.create_sheet('Legend')

# Add legend horizontally:
# A1: 'PK:', B1 to I1: Color-coded systems, J1: 'FK:', K1: Foreign Key color coding
legend_ws['A1'] = 'PK:'
legend_ws['J1'] = 'FK:'
legend_ws['A1'].font = pk_bold_font
legend_ws['J1'].font = pk_bold_font

# Add system colors to B1 to I1 and FK color to K1
legend_data = [
    ('D365', convert_to_argb('D9E1F2')),
    ('AXAPTA', convert_to_argb('C6E0B4')),
    ('SalesForce', convert_to_argb('FFF2CC')),
    ('PDM', convert_to_argb('F4CCCC')),
    ('Mobile Installer', convert_to_argb('D9D9D9')),
    ('Order Management', convert_to_argb('FFD966')),
    ('DWH (or unrecognized)', convert_to_argb('EAD1DC')),
    ('Budget.xlsx', convert_to_argb('D0E0E3'))
]

# Set system names and colors from B1 to I1
for i, (system, color) in enumerate(legend_data, start=2):
    legend_cell = legend_ws.cell(row=1, column=i, value=system)
    legend_cell.font = pk_bold_font  # Bold system names
    legend_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# Set FK color in K1
fk_legend_cell = legend_ws.cell(row=1, column=11, value='Foreign Key')
fk_legend_cell.font = pk_bold_font
fk_legend_cell.fill = fk_fill

# Save the workbook with PK/FK highlights and the legend
wb.save(output_file)

# Clear the screen
os.system('cls' if os.name == 'nt' else 'clear')

# Display the peanut ASCII art using raw string
peanut_art = r"""
  ,-~~-.___.
 / |  '     \         It was a dark and stormy night....
(  )         0              
 \_/-, ,----'            
    ====           //                     
   /  \-'~;    /~~~(O)
  /  __/~|   /       |     
=(  _____| (_________|   <3
"""

print(peanut_art)

print(f"Entity column names with PK/FK highlights and hyperlinks were saved to: {output_file}")
